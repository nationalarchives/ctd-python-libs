# Functions related to reading and writing data from spreadsheets

import os, re, shutil, csv
from openpyxl import Workbook, load_workbook
from openpyxl.styles import Font
from pathlib import Path

def getSpreadsheetValues(filename):
    ''' Gets spreadsheet by name and returns the spreadsheet as a worksheet and a list of column headings '''
    #path = os.path.join('data', filename) 
    wb = load_workbook(filename)
    
    sheet = wb.worksheets[0]
    values={}
    
    for col in sheet.columns:
        #column = [cell.value for cell in col if cell.value is not None]
        column = [cell.value if cell.value is not None else "" for cell in col]
        
        if len(column) > 0 and column.count("") != len(column):
            values[str(column[0]).strip()] = column[1:]
            
    return (values)


def getFileList(myDir):
    ''' Get a list of xlsx files in the given directory '''
    return [file for file in myDir.glob("[!~.]*.xlsx")]


def createSpreadsheetWithValues(path, filename, filenameExtra, values, newValues, filteredColumn, min):
    ''' print out a new spreadsheet with the supplied values (columns in newValues with the same title replace the original version in values, can also use filter columns to replace values within a column rather than an entire column)'''
    
    wb = Workbook()
    newSheet = wb.active
    
    col = 1
    
    #print(values)

    #print(newValues)
    
    for title, column in values.items():
        newSheet.cell(1, col, title).font = Font(bold=True)
    
        row = 2
        
        if title in newValues.keys():
            column = newValues[title]

        filteredColumn = zip(column, filter)           

        for filteredRow in filteredColumn:
            #print(str(row[1]) + ": " + str(x) + ", " + str(y))
            if (min and filteredRow[1]) or not min:
                newSheet.cell(row, col, filteredRow[0])
                row+=1
            
        col+=1 
    
    if row > 2:   
        if not os.path.exists(path):
            os.makedirs(path)

        newFilename = os.path.splitext(os.path.basename(filename))[0] + "_" + str(filenameExtra) + os.path.splitext(os.path.basename(filename))[1]
        newFile = os.path.join(path, newFilename)  
        wb.save(newFile)